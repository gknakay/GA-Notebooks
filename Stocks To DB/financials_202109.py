# -*- coding: utf-8 -*-
"""
Created on Fri Jan 28 18:34:38 2022

@author: Gökhan Akay
"""

import requests
from bs4 import BeautifulSoup

import pandas as pd
import xlrd
import os
from tqdm import tqdm

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def pull_financials():

    current_period = '2021/9'
    tam_directory = 'C:\\Users\\Gökhan Akay\\Desktop\\Google Drive\\TAM_202112.xlsx'
    
    workbook = xlrd.open_workbook(tam_directory)
    worksheet = workbook.sheet_by_name('Main')
    
    #excel'deki main sheet'indeki hisseleri aldık.
    stock_list = []
    for i in range(1,100):
        stock_list.append(worksheet.cell(i, 23).value)
        
    value_list = []
    for i in range(1,100):
        value_list.append(worksheet.cell(i, 42).value)
    
    stock_list = [stock for stock in stock_list if stock != '']
    value_list = [value for value in value_list if value != '']

    
# =============================================================================
#     # burda zaten çektiğimiz hisseleri buluyoruz.
#     os.chdir('C:\\Users\\Gökhan Akay\\Desktop\\BeautifulSoup\\Financials_202112')
#     current_stocks = [file[:-5] for file in os.listdir() if file.endswith('.xlsx')]
#     
#     #burda zaten çektiklerimizi eliyoruz. 
#     stock_list = sorted(list(set(stock_list).difference(current_stocks)))
#     
#     #burda kalan hisselerin güncel periyotlarını buluyoruz.
#     latest_quarters = []
#     #for i in range(0, len(stock_list)):
#     for i in tqdm (range(0,len(stock_list)), desc="Loading…", ascii=False, ncols=75):
#         stock = stock_list[i]
#         html_text = requests.get('https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse='+stock).text
#     
#         soup = BeautifulSoup(html_text, 'lxml')
#         table = soup.find('tbody', id = 'tbodyMTablo')
#         trs = table.find_all('tr')
#     
#     
#     
#         quarter_list = []
#         for i in range(0,4):
#             quarter_list.append(soup.find_all('th', class_ = 'form-group')[0].find_all('option')[i].text)
#     
#         latest_quarters.append(quarter_list[0])
# =============================================================================
    
    
    #burda son periyodu bizim yukarıda belirttiğimiz stockları buluyoruz. 
    stock_period = pd.DataFrame(zip(stock_list,value_list))
    stock_period.columns = ['stock','latest_period']
    stock_period.set_index('stock', inplace = True)
    stocks_to_pull = list(stock_period[stock_period.latest_period == 0].index)
    
    print("These stocks will be pulled:", stocks_to_pull)
    
    for i in range(0,len(stocks_to_pull)):
        
        stock = stocks_to_pull[i]
        html_text = requests.get('https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse='+stock).text
    
        soup = BeautifulSoup(html_text, 'lxml')
        table = soup.find('tbody', id = 'tbodyMTablo')
        trs = table.find_all('tr')
        
        quarter_list = []
        for i in range(0,4):
            quarter_list.append(soup.find_all('th', class_ = 'form-group')[0].find_all('option')[i].text)
        
        val_list = []
        for i in range(1,len(trs)):
            val_list.append(trs[i].find_all('td'))
            
            
        col_list = []
        for i in range(0,len(val_list)):
            col_list.append(val_list[i][0].text)
            
        col_list[0]
        col_list.remove('Gelir Tablosu')
        try:
            col_list.remove('Dipnot')
        except:
            pass
        
        
        try:
            col_list.remove('Nakit Akım Tablosu')
        except:
            pass
        
        
            
        q1 = []    
        for i in range(0,len(val_list)):
            try:
                q1.append(val_list[i][1].text)
            except:
                i = i+1
                
        q1[0]
        
        q2 = []    
        for i in range(0,len(val_list)):
            try:
                q2.append(val_list[i][2].text)
            except:
                i = i+1
        
        q2[0]
        
        
        q3 = []    
        for i in range(0,len(val_list)):
            try:
                q3.append(val_list[i][3].text)
            except:
                i = i+1
        
        q3[0]
        
        q4= []    
        for i in range(0,len(val_list)):
            try:
                q4.append(val_list[i][4].text)
            except:
                i = i+1   
            
        q4[0]
        
        q1 = [w.replace('.', '') for w in q1]
        q1 = list(map(float, q1))
    
        q2 = [w.replace('.', '') for w in q2]
        q2 = list(map(float, q2))
    
        q3 = [w.replace('.', '') for w in q3]
        q3 = list(map(float, q3))
    
        q4 = [w.replace('.', '') for w in q4]
        q4 = list(map(float, q4))
        
        col_list.insert(0,'')
        q1.insert(0,'')
        q2.insert(0,'')
        q3.insert(0,'')
        q4.insert(0,'')
        
        col_list.insert(72,'')
        q1.insert(72,'')
        q2.insert(72,'')
        q3.insert(72,'')
        q4.insert(72,'')
        
        col_list.insert(115,'')
        q1.insert(115,'')
        q2.insert(115,'')
        q3.insert(115,'')
        q4.insert(115,'')
        
        
        df = pd.DataFrame(index = col_list)
        df['q1'] = q1
        df['q2'] = q2
        df['q3'] = q3
        df['q4'] = q4
        
        df.columns = quarter_list
        
        df = df.iloc[0:124,:]
        
        df.to_excel(stock+'.xlsx')
        
        print(stock+' is pulled')
        
        
# =============================================================================
#         
#         tam = load_workbook(tam_directory)
#         worksheet = tam[stock]
#         #worksheet.insert_cols(3)
#         #C3:C125 --> U3:U125
#         worksheet.move_range('C1:U130', rows = 0, cols = 1)
#     
#         for row in range(3,126):
#             for col in range(3,4):
#                 #char = chr(65+col)
#                 char = get_column_letter(col)
#                 worksheet[char + str(row)] = df.iloc[row-2, col-3]
#           
#         # buraya dikkat, ilk önce ARCLK geldi onun formülasyonu doğru. 1 tane düzgün formüllü bir sheet lazım!    
#         for row in range(131,199):
#                 for col in range(3,23):
#                     #char = chr(65+col)
#                     char = get_column_letter(col)
#                     worksheet = tam['ARCLK']
#                     tam[stock][char + str(row)] = tam['ARCLK'][char + str(row)].value
#         
#         tam.save(tam_directory)
# 
# =============================================================================
    

if __name__ == '__main__':
    pull_financials()
