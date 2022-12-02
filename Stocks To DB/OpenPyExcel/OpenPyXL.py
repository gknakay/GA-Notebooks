# -*- coding: utf-8 -*-
"""
Created on Fri Feb  4 10:18:43 2022

@author: Gökhan Akay
"""

import openpyxl

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

workbook = load_workbook('deneme.xlsx')
worksheet = workbook.active

print(worksheet)
print(worksheet['C21'].value)
worksheet['C21'] = 'TEB ARF'

workbook.save('deneme2.xlsx')

print(workbook.sheetnames)

workbook['Sheet1']

workbook.create_sheet('İkinci Sheet')
print(workbook.sheetnames)


#adding data
wb = Workbook()
ws = wb.active
ws.title = 'Data'

# yan yana ekliyor. 
ws.append(['A','B','C','D'])
ws.append([1,2,3,4])
wb.save('deneme2.xlsx')

wb.create_sheet('Loop')

ws = wb['Data']

for row in range(1,11):
    for col in range(1,5):
        #char = chr(65+col)
        char = get_column_letter(col)
        print(ws[char + str(row)].value)
        
        

#inserting rows        
#bu şekilde diyince tam 20. satır boş olmuş oluyor.
worksheet.insert_rows(20)        
workbook.save('deneme.xlsx')
worksheet.insert_cols(2)        
workbook.save('deneme.xlsx')

