# -*- coding: utf-8 -*-
"""
Created on Fri Jan 28 18:34:38 2022

@author: Gökhan Akay
"""

import requests
from bs4 import BeautifulSoup

import pandas as pd
import xlrd
import os
from tqdm import tqdm

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

import yfinance as yf
from datetime import date, datetime



def pull_financials():

    current_period = '2021/12'
    examplary_stock = 'ARCLK' # malesef ilk önce bir hisseyi kendimiz çekmek durumundayız. Openpyxl kütüphanesi excel
                              # formüllerini otomatik olarak kaydıramıyor, text olarak algılıyor. 
    
    # temel analiz modülünün bulunduğu directory. 
    tam_directory = 'C:\\Users\\Gökhan Akay\\Desktop\\Google Drive\\TAM_202112.xlsx'
    
    workbook = xlrd.open_workbook(tam_directory)
    worksheet = workbook.sheet_by_name('Main')
    
    #excel'deki main sheet'indeki hisseleri aldık. Burası hisse listemiz.
    stock_list = []
    for i in range(1,100):
        stock_list.append(worksheet.cell(i, 23).value)
    
    stock_list = [stock for stock in stock_list if stock != '']
    
    # halihazırda çektiğimiz hisseleri buluyoruz.
    os.chdir('C:\\Users\\Gökhan Akay\\Desktop\\BeautifulSoup\\Financials_202112')
    current_stocks = [file[:-5] for file in os.listdir() if file.endswith('.xlsx')]
    
    # halihazırda çektiklerimizi eliyoruz. 
    stock_list = sorted(list(set(stock_list).difference(current_stocks)))
    
    # kalan hisselerin güncel periyotlarını buluyoruz.
    # her bir hisse için işyatırıma gidip maksimum periyodu getirtiyoruz.
    latest_quarters = []
    print("Latest quarters are being checked...")
    for i in tqdm (range(0,len(stock_list)), desc="Loading…", ascii=False, ncols=75):
        stock = stock_list[i]
        html_text = requests.get('https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse='+stock).text
    
        soup = BeautifulSoup(html_text, 'lxml')
        table = soup.find('tbody', id = 'tbodyMTablo')
        trs = table.find_all('tr')    
    
        quarter_list = []
        for i in range(0,4):
            quarter_list.append(soup.find_all('th', class_ = 'form-group')[0].find_all('option')[i].text)
    
        latest_quarters.append(quarter_list[0])
    
    
    #burda son periyodu bizim yukarıda belirttiğimiz periyot olan hisseleri buluyoruz.
    stock_period = pd.DataFrame(zip(stock_list,latest_quarters))
    stock_period.columns = ['stock','latest_period']
    stock_period.set_index('stock', inplace = True)
    stocks_to_pull = list(stock_period[stock_period.latest_period == current_period].index)    
    if len(stocks_to_pull) == 0:
        print("There is no stocks to pull. Program is terminating.")
    else:
        print("These stocks will be pulled:", stocks_to_pull)
    
    # çekilecek her bir hisse için döngüye giriyoruz.
    for i in range(0,len(stocks_to_pull)):
        
        stock = stocks_to_pull[i]
        html_text = requests.get('https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse='+stock).text
    
        soup = BeautifulSoup(html_text, 'lxml')
        table = soup.find('tbody', id = 'tbodyMTablo')
        trs = table.find_all('tr')
        
        quarter_list = []
        for i in range(0,4):
            quarter_list.append(soup.find_all('th', class_ = 'form-group')[0].find_all('option')[i].text)
        
        val_list = []
        for i in range(1,len(trs)):
            val_list.append(trs[i].find_all('td'))
            
            
        col_list = []
        for i in range(0,len(val_list)):
            col_list.append(val_list[i][0].text)
            
        col_list[0]
        col_list.remove('Gelir Tablosu')
        try:
            col_list.remove('Dipnot')
        except:
            pass
        
        
        try:
            col_list.remove('Nakit Akım Tablosu')
        except:
            pass
        
        
            
        q1 = []    
        for i in range(0,len(val_list)):
            try:
                q1.append(val_list[i][1].text)
            except:
                i = i+1
                
        q1[0]
        
        q2 = []    
        for i in range(0,len(val_list)):
            try:
                q2.append(val_list[i][2].text)
            except:
                i = i+1
        
        q2[0]
        
        
        q3 = []    
        for i in range(0,len(val_list)):
            try:
                q3.append(val_list[i][3].text)
            except:
                i = i+1
        
        q3[0]
        
        q4= []    
        for i in range(0,len(val_list)):
            try:
                q4.append(val_list[i][4].text)
            except:
                i = i+1   
            
        q4[0]
        
        q1 = [w.replace('.', '') for w in q1]
        q1 = list(map(float, q1))
    
        q2 = [w.replace('.', '') for w in q2]
        q2 = list(map(float, q2))
    
        q3 = [w.replace('.', '') for w in q3]
        q3 = list(map(float, q3))
    
        q4 = [w.replace('.', '') for w in q4]
        q4 = list(map(float, q4))
        
        col_list.insert(0,'')
        q1.insert(0,'')
        q2.insert(0,'')
        q3.insert(0,'')
        q4.insert(0,'')
        
        col_list.insert(72,'')
        q1.insert(72,'')
        q2.insert(72,'')
        q3.insert(72,'')
        q4.insert(72,'')
        
        col_list.insert(115,'')
        q1.insert(115,'')
        q2.insert(115,'')
        q3.insert(115,'')
        q4.insert(115,'')
        
        
        df = pd.DataFrame(index = col_list)
        df['q1'] = q1
        df['q2'] = q2
        df['q3'] = q3
        df['q4'] = q4
        
        df.columns = quarter_list
        
        # gelir tablosu ve dipnot'tan sonra gelen kısmı eliyoruz. 
        df = df.iloc[0:124,:]
        
        #excel'e yazdırıyoruz. 
        df.to_excel(stock+'.xlsx')
        
        print(stock+' is pulled')
        
        
        # buradan sonra çektiğimiz datayı temel analiz modülüne yazdırıyoruz.
        
        tam = load_workbook(tam_directory)
        worksheet = tam[stock]
        #worksheet.insert_cols(3)
        #C3:C125 --> U3:U125


        # hissenin sheet'ine gidip yeni periyot için yer açıyoruz, 1 kolon sağa kaydırıyoruz. 
        worksheet.move_range('C1:U130', rows = 0, cols = 1)
    
        # burası bilanço kısmını dolduruyor. 125. satıra kadar olan bölüm.
        for row in range(3,126):
            for col in range(3,4):
                #char = chr(65+col)
                char = get_column_letter(col)
                worksheet[char + str(row)] = df.iloc[row-2, col-3]
          
        # buraya dikkat, ilk önce ARCLK geldi onun formülasyonu doğru. 1 tane düzgün formüllü bir sheet lazım!  
        # burada bilanço metrikleri hesaplanıyor. 
        
        for row in range(131,199):
                for col in range(3,23):
                    #char = chr(65+col)
                    char = get_column_letter(col)
                    worksheet = tam[examplary_stock]
                    tam[stock][char + str(row)] = tam[examplary_stock][char + str(row)].value
        
        # bilanço gününden 1 gün önceki fiyatı yapıştırıyoruz. 
        today = date.today().strftime("%Y-%m-%d")
        data = pd.DataFrame(yf.download(stock+'.IS','2022-01-01',today)['Close'])
        price = data['Close'][-1]
        tam[stock]['C130'] = price
        
        # temel analiz modülü yapılan değişiklikler sonrası kayıt ediliyor. 
        tam.save(tam_directory)

    

if __name__ == '__main__':
    pull_financials()

