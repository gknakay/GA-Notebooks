# -*- coding: utf-8 -*-
"""
Created on Fri Feb  4 12:03:01 2022

@author: Gökhan Akay
"""

import requests
from bs4 import BeautifulSoup

import pandas as pd
import xlrd
import os
from tqdm import tqdm

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


stock_list = [file[:-5] for file in os.listdir() if file.endswith('.xlsx')]
stock_list.remove('ULUSE')
stock_list = [stock.replace("'", "") for stock in stock_list]

tam_directory = 'C:\\Users\\Gökhan Akay\\Desktop\\Google Drive\\TAM_202112.xlsx'
stock_directory = 'C:\\Users\\Gökhan Akay\\Desktop\\BeautifulSoup\\Financials_202109'




for stock in stock_list:
    
    workbook = load_workbook(stock_directory+'\\'+stock+'.xlsx')
    worksheet = workbook.active
   
    temp_list = []
    for row in range(3,126):
        char = get_column_letter(2)
        temp_list.append(worksheet[char + str(row)].value)
    
    workbook = load_workbook(tam_directory)
    worksheet = workbook[stock]
    
    worksheet.move_range('C1:U130', rows = 0, cols = 1)
    for row in range(3,126):
            for col in range(3,4):
                #char = chr(65+col)
                char = get_column_letter(col)
                worksheet[char + str(row)] = temp_list[row-3]
          
        # buraya dikkat, ilk önce ARCLK geldi onun formülasyonu doğru. 1 tane düzgün formüllü bir sheet lazım!    
    for row in range(131,199):
            for col in range(3,23):
                #char = chr(65+col)
                char = get_column_letter(col)
                worksheet = workbook['ASELS']
                workbook[stock][char + str(row)] = workbook['ASELS'][char + str(row)].value
    
    workbook.save(tam_directory)
    print(stock+' is saved to TAM')
    
